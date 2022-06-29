import sys
from datetime import timezone, timedelta, datetime

import openpyxl.utils.exceptions
from dateutil import parser
from openpyxl import load_workbook
import xlrd
import xlwings as xw

import very_general_functions
from very_general_functions import is_str_num, is_str_float
from time_func import adjust_year, from_dt2str_yr_day

vol='/users/kylefoley/'

xl_w_macro = f'{vol}documents/codes/excel/python_macros2.xlsm'
get_sheet_read = lambda wb, name: wb.sheet_by_name(name)
en = enumerate
p = print


def get_from_excel_date(ws, rw, col, debug=False):
    obj = get_from_excel(ws, rw, col, debug)
    if isinstance(obj, str):
        try:
            dt = parser.parse(obj)
            return dt
        except:
            return obj
    else:
        return obj


def get_date_xlm(str_date, last_dt=[], tyear=0):
    skip_errors = 0
    if skip_errors == 0:
        c = xlrd.xldate.xldate_as_datetime(str_date, 0)
        if tyear != 0:
            return adjust_year(c, tyear, skip_errors)
        else:
            return c

    else:

        b = datetime.now()
        try:
            c = xlrd.xldate.xldate_as_datetime(str_date, 0)

            if not isinstance(c, datetime.datetime):
                if last_dt == []:
                    p('not a datetime object')
                    return b
                else:
                    p('error')
                    p(from_dt2str_yr_day(last_dt))
                    return last_dt
            elif tyear != 0:
                return adjust_year(c, tyear)
            else:
                return c
        except:
            p('not a datetime object')
            p('error')
            if isinstance(last_dt, datetime.datetime):
                p(from_dt2str_yr_day(last_dt))
                return last_dt
            else:
                return b


def get_last_row(worksheet, col_num, row_num=500_000):
    str3 = worksheet.cell(row=row_num, column=col_num).value
    while not str3:
        str3 = worksheet.cell(row=row_num, column=col_num).value
        row_num -= 1

    return row_num + 1


def get_lrow_of_sh(worksheet, start_row=5000):
    lst_col = get_last_col(worksheet, 1)
    highest = 0
    for col_num in range(lst_col, 1, -1):
        max_row = start_row
        str3 = worksheet.cell(row=max_row, column=col_num).value
        while not str3:
            str3 = worksheet.cell(row=max_row, column=col_num).value
            max_row -= 1
            if max_row < 3: break
        if max_row > highest:
            highest = max_row

    return highest + 1


def get_last_col(worksheet, row_num, col_num=500):
    str3 = worksheet.cell(row=row_num, column=col_num).value
    while not str3 and col_num:
        str3 = worksheet.cell(row=row_num, column=col_num).value
        col_num -= 1
    return col_num + 2

def get_last_col_read(worksheet, row_num, col_num=500):
    str3 = 0
    while not str3 and col_num:
        try:
            str3 = worksheet.cell(row_num, col_num).value
            col_num -= 1
        except:
            col_num -= 1
            str3 = 0
    return col_num + 2


def get_last_rowm(ws, rw, col):
    for tm_last_row in range(rw, 1, -1):
        val = get_from_excelm(ws, tm_last_row, col)
        if val != "":
            if tm_last_row == 2:
                raise Exception
            return tm_last_row


def get_target_colm(ws, rw, target):
    for col in range(0, 200):
        val = get_from_excelm(ws, rw, col)
        if val == target:
            return col
    else:
        raise Exception


def make_cell_red(ws, row, col):
    xls_cell = ws.cell(row=row, column=col)
    xls_cell.font = xls_cell.font.copy(color='FFFF0000')


def get_target_col(ws, rw, target):
    for col in range(1, 200):
        val = get_from_excel(ws, rw, col)
        if val == target:
            return col
    else:
        raise Exception


def get_target_rw(ws, col, target, start, stop):
    for rw in range(start, stop):
        val = get_from_excel(ws, rw, col)
        if val == target:
            return rw
    else:
        raise Exception


def get_target_rwm(ws, col, target, start, stop):
    for rw in range(start, stop):
        val = get_from_excelm(ws, rw, col)
        if val == target:
            return rw
    else:
        raise Exception


def load_workbook_read(file):
    return xlrd.open_workbook(file)


def load_workbook_write(file):
    return load_workbook(file)


def fromstr2int(lst):
    for e, x in en(lst):
        for f, y in en(x):
            if e == 431 and f == 5:
                bb = 8

            if y in ['01', '00', '10']:
                pass
            elif is_str_num(y):
                lst[e][f] = int(y)
            elif is_str_float(y):
                lst[e][f] = float(y)

    return lst

def from_sheet_tpl0(ws, col=1, stripp=1):
    '''
    assumes that row 1 has the largest amount of columns
    '''
    lrow = get_lrow_of_sh(ws)
    lcol = get_last_col(ws,col)
    return from_sheet_tpl(ws, 1, lrow+1, lcol, stripp)


def from_sheet_tpl(ws, start=0, stop=0, cstop=0, stripp=1):
    if not start:
        cstop = get_last_col(ws,0)
        cstop2 = get_last_col(ws,1)
        if cstop2 > cstop:
            cstop = cstop2
        start = get_last_row(ws,0)
        start2 = get_last_row(ws,1)
        if start2 > start:
            start = start2



    lst = []
    cols = [None] * cstop
    lst.append(cols)
    for rw in range(start, stop):
        cols = [None] * cstop
        for col in range(1, cstop):
            val = get_from_excel(ws, rw, col,0,stripp)
            cols[col] = val
        lst.append(cols)
    return lst


def from_lst2sheet(ws, lst, **dct):
    for e, y in en(lst):
        very_general_functions.print_intervals(e,100,None,len(lst))
        for f, x in en(y):
            if e > 0 and f > 0:
                put_into_excel(ws, e, f, lst[e][f],**dct)
    return

def convert19002dt(num):
    num = int(num)  - 2
    dt = datetime(year=1900,month=1,day=1)
    dt = dt + timedelta(days=num)
    return from_dt2str_yr_day(dt)



def from_sheet_tpl_read(ws, cstart=0, cstop=0, convert2none=True, all_str=False):
    if not cstart:
        cstop = get_last_col_read(ws,0)
        cstop2 = get_last_col_read(ws,1)
        if cstop2 > cstop:
            cstop = cstop2
        # start = get_last_row(ws,0)
        # start2 = get_last_row(ws,1)
        # if start2 > start:
        #     start = start2

    lst = ws._cell_values
    cols = [None] * ((cstop - cstart) + 1)
    lst.insert(0, cols)
    for e, lst1 in en(lst):
        lst1.insert(0, "")
        if all_str:
            lst[e] = [make_str(x) for x in lst1]
        else:
            lst[e] = lst1


    if convert2none:
        return change_to_none(lst)
    else:
        return lst


def make_str(x):
    try:
        key = str(int(x))
    except:
        key = str(x)
    return key

def from_lst2book(lst, dest):
    wb = load_workbook_write(f"{vol}documents/codes/excel/blank.xlsx")
    sh = get_sheet(wb,'Sheet1')
    # sh = get_sheet_read(wb,'Sheet1')
    for x in lst:
        if not type(x) == list:
            p ('must be a list of list to transfer to excel')
            assert 0
    lst.insert(0,['',''])
    lst = [[''] + x for x in lst]
    bad = set()
    from_lst2sheet(sh,lst,{'test':1,'bad':bad})
    wb.save(dest+'.xlsx')
    # open_wb(dest+'.xlsx')

def from_book2lst(path):
    if not path.endswith('.xlsx'):
        path += '.xlsx'

    wb = load_workbook_read(path)
    lsts = []
    for sh in wb._sheet_names:
        sh = get_sheet_read(wb, sh)
        lst = from_sheet_tpl_read(sh)
        lsts.append(lst)
    return lsts

def change_to_none(lst):
    for f, x in en(lst):
        for e, y in en(x):
            if isinstance(y, str):
                lst[f][e] = lst[f][e].strip()
                y = lst[f][e]
            if y == "":
                lst[f][e] = None
    return lst


def put_into_excel(worksheet, rw, col, val, **kwargs):
    if val == None:
        val = ""
    if val and type(val) == str and len(val) == 1 and ord(val) == 1:
        val = ""

    if str(val).startswith('tree; tree trunk; mast; oar;'):
        bb = 8
    bold = kwargs.get("bold")
    size = kwargs.get("size", 12)
    test = kwargs.get("test")
    bad = kwargs.get("bad")

    # if workbook:
    #     put_excel_ud(workbook, worksheet, rw, col, val)
    # else:
    if type(val) == str:
        for i in range(1,33):
            val = val.replace(chr(i),' ')

    try:
        worksheet.cell(row=rw, column=col).value = val
    except openpyxl.utils.exceptions.IllegalCharacterError:
        # if test:
        #     for x in val:
        #         try:
        #             worksheet.cell(row=rw, column=col).value = x
        #         except:
        #             bad.add(ord(x))
        p (f'bad character in {val}')

        return

    if bold:
        make_bold(worksheet, rw, col)
    if size != 12:
        adjust_size(worksheet, rw, col, size)
    return


def make_bold(ws, rw, col):
    b = ws.cell(row=rw, column=col)
    b.font = b.font.copy(bold=True)


def adjust_size(ws, rw, col, font_size):
    b = ws.cell(row=rw, column=col)
    b.font = b.font.copy(size=font_size)

def get_headers_xlx(ws):
    dct = {}
    b = 1
    while 1:
        str1 = ws.cell(row=1,column=b).value
        if str1:
            dct[str1] = b
        else:
            break
        b += 1
    return dct



def put_excel_ud(workbook, worksheet, rw, col, val):
    wb = xw.Book(xl_w_macro)
    b = wb.macro('input_from_python')
    b.run(workbook, worksheet, rw, col, val)
    return


def close_wb(workbook):
    wb = xw.Book(xl_w_macro)
    # c = wb.macro('update_sheet')
    # c.run(workbook)
    b = wb.macro('close_wb')
    b.run(workbook)


def open_wb(workbook, extra_path=""):
    if not extra_path:
        str1 = workbook
    else:
        str1 = extra_path + workbook
    wb = xw.Book(xl_w_macro)
    b = wb.macro('open_wb')
    if str1[0] == '/': str1 = str1[1:]
    str1 = str1.replace("/", ":")
    b.run(str1)


def save_excel_sheet(workbook):
    wb = xw.Book(xl_w_macro)
    b = wb.macro('save_workbook')
    b.run(workbook)


def activate_time4():
    wb = xw.Book(xl_w_macro)
    b = wb.macro('activate_time4')
    b.run()


def disable_screen():
    wb = xw.Book(xl_w_macro)
    b = wb.macro('disable_screen')
    b.run()


def enable_screen():
    wb = xw.Book(xl_w_macro)
    b = wb.macro('enable_screen')
    b.run()


def save_all_books():
    wb = xw.Book(xl_w_macro)
    b = wb.macro('save_all_books')
    b.run()


def change_date_system(workbook):
    wb = xw.Book(xl_w_macro)
    b = wb.macro('date_system')
    b.run(workbook)


def savexl_reld(workbook, worksheet, extra_path):
    save_excel_sheet(workbook)
    wrkbook = load_workbook(extra_path + workbook)
    sht = get_sheet(wrkbook, worksheet)
    assert sht != None
    return sht


def update_excel_screen(workbook):
    wb = xw.Book(xl_w_macro)
    b = wb.macro('update_sheet')
    b.run(workbook)


def change_symbols():
    wb = xw.Book(xl_w_macro)
    b = wb.macro('logical_symbols_dict')
    b.run()


def use_list(sheet, rw, col):
    return sheet[rw][col]


def get_from_excel(worksheet, rw, col, debug=False, strip=1):
    if debug:
        return use_list(worksheet, rw, col)

    val = worksheet.cell(row=rw, column=col).value
    if isinstance(val, str) and strip:
        return val.strip()
    else:
        return val


def get_from_excel2(worksheet, rw, col):
    if worksheet.cell(row=rw, column=col).value == None: return 0
    return worksheet.cell(row=rw, column=col).value


def put_into_excelm(worksheet, rw, col, val):
    try:
        worksheet._cell_values[rw][col] = val
    except:
        pass
    return


def get_from_excelm(worksheet, rw, col):
    try:
        return worksheet._cell_values[rw][col]
    except:
        return ""


def get_xl_categories(sheet):
    dict1 = {}
    for col in range(300, 1, -1):
        str1 = get_from_excel(sheet, 1, col)
        if str1:
            dict1[str1] = col
    return dict1


def get_xl_categories_lst(lst):
    return {lst[1][col]: col for col in range(len(lst[1])) if lst[1][col]}


def get_xl_categoriesm(sheet, i=0):
    str1 = get_from_excelm(sheet, 0, i)
    dict1 = {}
    while str1 != "":
        dict1[str1] = i
        i += 1
        str1 = get_from_excelm(sheet, 0, i)

    return dict1


def get_sheet(workbk, name):
    for sheet in workbk.worksheets:
        if sheet.title == name:
            return sheet
    p(f"sheet {name} does not exist in the workbook")
    raise Exception


def convert_date(cell_date):
    if cell_date == None: return None
    if isinstance(cell_date, str): return cell_date
    return str(cell_date.month) + "/" + str(cell_date.day) + "/" + str(cell_date.year)[2:]


def sum_date(date, sheet, cell_date2, col, rw, tnum):
    cell_date = get_from_excel(sheet, rw, 1)
    cell_date2 = convert_date(cell_date)

    while date == cell_date2:
        num = get_from_excel(sheet, rw, col)
        if num == None: num = 0
        tnum += num
        rw -= 1
        cell_date = get_from_excel(sheet, rw, 1)
        cell_date2 = convert_date(cell_date)
    return tnum
